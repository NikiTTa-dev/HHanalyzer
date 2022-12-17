from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import csv
import datetime


class Vacancy:
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, vac):
        self.name = vac['name']
        salary_from = int(float("".join(vac['salary_from'].split())))
        salary_to = int(float("".join(vac['salary_to'].split())))
        self.salary = (salary_from + salary_to) * self.currency_to_rub[vac['salary_currency']] // 2
        self.area_name = vac['area_name']
        self.published_at = datetime.datetime.strptime(vac['published_at'], '%Y-%m-%dT%H:%M:%S%z')


class DataSet:
    def __init__(self, file_name: str, vacancies_objects: list):
        self.file_name = file_name
        self.vacancies_objects = vacancies_objects

    def fill_vacancies(self):
        vacancies, list_naming = self.read_csv()
        self.vacancies_objects = self.csv_filer(vacancies, list_naming)

    def read_csv(self):
        list_naming = []
        vacancies = []
        with open(self.file_name, encoding='utf-8-sig') as file:
            file_reader = csv.reader(file, delimiter=",")
            flag = True
            for row in file_reader:
                if flag:
                    flag = not flag
                    list_naming = row
                else:
                    if "" in row or len(row) != len(list_naming):
                        continue
                    vacancies.append(row)

        return vacancies, list_naming

    def csv_filer(self, vacs, list_naming):
        vacancies = list()
        for row in vacs:
            current = {}
            for i in range(len(row)):
                current[list_naming[i]] = row[i]
            vacancies.append(Vacancy(current))
        return vacancies


class MyTuple:
    def __init__(self, salary: int, count: int):
        self.totalSalary = salary
        self.count = count

    totalSalary = 0
    count = 0


class InputConect:
    years = {
    }

    cities = {
    }

    vacancies = {
    }

    file_name = ""
    profession = ""
    city_count = 0

    def start_input(self):
        self.file_name = input('Введите название файла: ')
        self.profession = input('Введите название профессии: ')
        self.city_count = 0

    def count_vacancies(self, vacancies: list):
        for vacancy in vacancies:
            self.city_count += 1
            year = int(vacancy.published_at.year)
            if year not in self.years.keys():
                self.years[year] = MyTuple(vacancy.salary, 1)
                self.vacancies[year] = MyTuple(0, 0)
            else:
                self.years[year].totalSalary += vacancy.salary
                self.years[year].count += 1

            if vacancy.area_name not in self.cities.keys():
                self.cities[vacancy.area_name] = MyTuple(vacancy.salary, 1)
            else:
                self.cities[vacancy.area_name].totalSalary += vacancy.salary
                self.cities[vacancy.area_name].count += 1

            if self.profession in vacancy.name:
                self.vacancies[year].totalSalary += vacancy.salary
                self.vacancies[year].count += 1

    def normalize_statistic(self):
        for year in self.years.keys():
            self.years[year].totalSalary = int(self.years[year].totalSalary // self.years[year].count)

        delete = list()
        for city in self.cities.keys():
            percent_count = round(self.cities[city].count / self.city_count, 4)
            if percent_count < 0.01:
                delete.append(city)
            else:
                self.cities[city].totalSalary = int(self.cities[city].totalSalary // self.cities[city].count)
                self.cities[city].count = percent_count
        for city in delete:
            del [self.cities[city]]

        for year in self.vacancies.keys():
            if self.vacancies[year].count != 0:
                self.vacancies[year].totalSalary = int(self.vacancies[year].totalSalary // self.vacancies[year].count)

    def print_answer(self):
        self.print_one("Динамика уровня зарплат по годам:", self.years, "totalSalary")
        self.print_one("Динамика количества вакансий по годам:", self.years, "count")

        self.print_one("Динамика уровня зарплат по годам для выбранной профессии:", self.vacancies, "totalSalary")
        self.print_one("Динамика количества вакансий по годам для выбранной профессии:", self.vacancies, "count")

        cities_sorted = sorted(self.cities, key=lambda x: self.cities[x].totalSalary, reverse=True)
        del cities_sorted[10:]
        self.print_for_cities("Уровень зарплат по городам (в порядке убывания):", self.cities,
                              cities_sorted, "totalSalary")
        cities_sorted = sorted(self.cities, key=lambda x: self.cities[x].count, reverse=True)
        del cities_sorted[10:]
        self.print_for_cities("Доля вакансий по городам (в порядке убывания):", self.cities,
                              cities_sorted, "count")

    def print_for_cities(self, output: str, field: dict, names: list, value_name):
        flag = False
        print(output, end='')
        index = 0
        for name in names:
            if index == 0:
                print(' {', end='')
            printEnd = ', '
            if index == len(names) - 1:
                printEnd = ''
                flag = True
            print(f"'{name}': {getattr(field[name], value_name)}", end=printEnd)
            index += 1
        if flag:
            print('}')

    def print_one(self, output: str, field: dict, value_name: str):
        flag = False
        print(output, end='')
        index = 0
        for year in field.keys():
            if index == 0:
                print(' {', end='')
                index += 1
            printEnd = ', '
            if year == max(field.keys()):
                printEnd = ''
                flag = True
            print(f"{year}: {getattr(field[year], value_name)}", end=printEnd)
        if flag:
            print('}')


class Report:
    years_salary = {}
    years_count = {}
    years_salary_vac = {}
    years_count_vac = {}
    area_salary = {}
    area_count = {}

    def years_preparer(self, field: dict, value_name: str, dest: dict):
        for year in field.keys():
            dest[year] = getattr(field[year], value_name)

    def citites_preparer(self, field: dict, names: list, value_name: str, dest: dict):
        for name in names:
            dest[name] = getattr(field[name], value_name)

    def prepare_data(self, years: dict, vacancies: dict, cities: dict):
        self.years_preparer(years, "totalSalary", self.years_salary)
        self.years_preparer(years, "count", self.years_count)
        self.years_preparer(vacancies, "totalSalary", self.years_salary_vac)
        self.years_preparer(vacancies, "count", self.years_count_vac)

        cities_sorted = sorted(cities, key=lambda x: cities[x].totalSalary, reverse=True)
        del cities_sorted[10:]
        self.citites_preparer(cities, cities_sorted, "totalSalary", self.area_salary)

        cities_sorted = sorted(cities, key=lambda x: cities[x].count, reverse=True)
        del cities_sorted[10:]
        self.citites_preparer(cities, cities_sorted, "count", self.area_count)

    def as_text(self, value: str):
        if value is None:
            return ""
        return str(value)

    def generate_excel(self):
        wb = Workbook()
        del wb['Sheet']
        sheet = wb.create_sheet('Статистика по годам')
        thin = Side(border_style="thin", color="000000")

        sheet["A1"] = "Год"
        sheet["A1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["A1"].font = Font(bold=True)

        sheet["B1"] = "Средняя зарплата"
        sheet["B1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["B1"].font = Font(bold=True)

        sheet["C1"] = "Средняя зарплата - Программист"
        sheet["C1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["C1"].font = Font(bold=True)

        sheet["D1"] = "Количество вакансий"
        sheet["D1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["D1"].font = Font(bold=True)

        sheet["E1"] = "Количество вакансий - Программист"
        sheet["E1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["E1"].font = Font(bold=True)

        for row, (year, value) in enumerate(self.years_salary.items(), start=2):
            sheet[f"A{row}"] = year
            sheet[f"A{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"B{row}"] = value
            sheet[f"B{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"C{row}"] = self.years_salary_vac[year]
            sheet[f"C{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"D{row}"] = self.years_count[year]
            sheet[f"D{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"E{row}"] = self.years_count_vac[year]
            sheet[f"E{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for column_cells in sheet.columns:
            length = max(len(self.as_text(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        sheet = wb.create_sheet('Статистика по городам')
        sheet["A1"] = "Город"
        sheet["A1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["A1"].font = Font(bold=True)

        sheet["B1"] = "Уровень зарплат"
        sheet["B1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["B1"].font = Font(bold=True)

        sheet["C1"].font = Font(bold=True)

        sheet["D1"] = "Город"
        sheet["D1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["D1"].font = Font(bold=True)

        sheet["E1"] = "Доля вакансий"
        sheet["E1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["E1"].font = Font(bold=True)

        for row, (year, value) in enumerate(self.area_salary.items(), start=2):
            sheet[f"A{row}"] = year
            sheet[f"A{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"B{row}"] = value
            sheet[f"B{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for row, (year, value) in enumerate(self.area_count.items(), start=2):
            sheet[f"D{row}"] = year
            sheet[f"D{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"E{row}"] = value
            sheet[f"E{row}"].number_format = FORMAT_PERCENTAGE_00
            sheet[f"E{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for column_cells in sheet.columns:
            length = max(len(self.as_text(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        wb.save('report.xlsx')


inputer = InputConect()
inputer.start_input()
dataset = DataSet(inputer.file_name, list())
dataset.fill_vacancies()
inputer.count_vacancies(dataset.vacancies_objects)
inputer.normalize_statistic()
inputer.print_answer()
reporter = Report()
reporter.prepare_data(inputer.years, inputer.vacancies, inputer.cities)
reporter.generate_excel()



