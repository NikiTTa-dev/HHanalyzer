from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import csv
import matplotlib.pyplot as pyplot
import numpy as np
import datetime
from jinja2 import Environment, FileSystemLoader
import pdfkit


class Vacancy:
    """
    Класс для описания вакансии

    Atributes
    ---------
    name : str
        название професии
    salary : int
        зарплата
    area_name : str
        название региона
    published_at : str
        дата публикации
    currency_to_rub : dict
        словарь перевода валюты в рубли
    """
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, vac):
        """
        Инициализация объекта
        :param vac: dict
            вакансия
        """
        self.name = vac['name']
        salary_from = int(float("".join(vac['salary_from'].split())))
        salary_to = int(float("".join(vac['salary_to'].split())))
        self.salary = (salary_from + salary_to) * self.currency_to_rub[vac['salary_currency']] // 2
        self.area_name = vac['area_name']
        self.published_at = ".".join(reversed(vac["published_at"][:10].split("-")))
            #datetime.datetime.strptime(vac['published_at'], '%Y-%m-%dT%H:%M:%S%z')


class DataSet:
    """
    Читает csv и заполняет список вакансий

    Attributes
    ----------
    file_name: str
        имя файла
    vacancies_objects: list
        список вакансий
    """
    def __init__(self, file_name: str, vacancies_objects: list):
        """
        инициализация объекта
        :param file_name: str
            имя файла
        :param vacancies_objects: str
            список вакансий
        """
        self.file_name = file_name
        self.vacancies_objects = vacancies_objects

    def fill_vacancies(self):
        """
        Читает вакансии и фильтрует их
        """
        vacancies, list_naming = self.read_csv()
        self.vacancies_objects = self.csv_filer(vacancies, list_naming)

    def read_csv(self):
        """
        чтение csv файла
        :return: (str[], str[])
            значения(сами вакансии), заглавия(параметры)
        """
        list_naming = []
        vacancies = []
        with open(self.file_name, encoding='utf-8-sig') as file:
            file_reader = csv.reader(file, delimiter=",")
            flag = True
            for row in file_reader:
                if flag:
                    flag = not flag
                    list_naming = row
                else:
                    if "" in row or len(row) != len(list_naming):
                        continue
                    vacancies.append(row)

        return vacancies, list_naming

    def csv_filer(self, vacs, list_naming):
        """
        фильтрует вакансии
        :param vacs: list
            вакансии
        :param list_naming: list
            название полей
        :return: list
            итоговы список вакансий
        """
        vacancies = list()
        for row in vacs:
            current = {}
            for i in range(len(row)):
                current[list_naming[i]] = row[i]
            vacancies.append(Vacancy(current))
        return vacancies


class MyTuple:
    """
    Кастомный класс tuple

    Attributes
    ----------
    totalSalary: int
        зарплата
    count: int
        количество
    """
    def __init__(self, salary: int, count: int):
        """
        инициализация объекта
        :param salary: int
            зарплата
        :param count: int
            количество
        """
        self.totalSalary = salary
        self.count = count

    totalSalary = 0
    count = 0


class InputConect:
    """
    класс ввода/вывода

    Attributes
    ----------
    years: dict
        словарь лет
    cities: dict
        словарь городов
    vacancies: dict
        словарь вакансий
    file_name: str
        название файла
    profession: str
        название профессии
    city_count: int
        количество городов
    """
    years = {
    }

    cities = {
    }

    vacancies = {
    }

    file_name = ""
    profession = ""
    city_count = 0

    def start_input(self):
        """
        начинает пользовательский ввод
        """
        self.file_name = input('Введите название файла: ')
        self.profession = input('Введите название профессии: ')
        self.city_count = 0

    def count_vacancies(self, vacancies: list):
        """
        заполняет словари для вакансий
        :param vacancies: list
            вакансии
        """
        for vacancy in vacancies:
            self.city_count += 1
            year = int(vacancy.published_at.split(".")[2])
            if year not in self.years.keys():
                self.years[year] = MyTuple(vacancy.salary, 1)
                self.vacancies[year] = MyTuple(0, 0)
            else:
                self.years[year].totalSalary += vacancy.salary
                self.years[year].count += 1

            if vacancy.area_name not in self.cities.keys():
                self.cities[vacancy.area_name] = MyTuple(vacancy.salary, 1)
            else:
                self.cities[vacancy.area_name].totalSalary += vacancy.salary
                self.cities[vacancy.area_name].count += 1

            if self.profession in vacancy.name:
                self.vacancies[year].totalSalary += vacancy.salary
                self.vacancies[year].count += 1

    def normalize_statistic(self):
        """
        обрабатывает статистику
        """
        for year in self.years.keys():
            self.years[year].totalSalary = int(self.years[year].totalSalary // self.years[year].count)

        delete = list()
        for city in self.cities.keys():
            percent_count = round(self.cities[city].count / self.city_count, 4)
            if percent_count < 0.01:
                delete.append(city)
            else:
                self.cities[city].totalSalary = int(self.cities[city].totalSalary // self.cities[city].count)
                self.cities[city].count = percent_count
        for city in delete:
            del [self.cities[city]]

        for year in self.vacancies.keys():
            if self.vacancies[year].count != 0:
                self.vacancies[year].totalSalary = int(self.vacancies[year].totalSalary // self.vacancies[year].count)

    def print_answer(self):
        """
        Печать ответа в консоль
        """
        self.print_one("Динамика уровня зарплат по годам:", self.years, "totalSalary")
        self.print_one("Динамика количества вакансий по годам:", self.years, "count")

        self.print_one("Динамика уровня зарплат по годам для выбранной профессии:", self.vacancies, "totalSalary")
        self.print_one("Динамика количества вакансий по годам для выбранной профессии:", self.vacancies, "count")

        cities_sorted = sorted(self.cities, key=lambda x: self.cities[x].totalSalary, reverse=True)
        del cities_sorted[10:]
        self.print_for_cities("Уровень зарплат по городам (в порядке убывания):", self.cities,
                              cities_sorted, "totalSalary")
        cities_sorted = sorted(self.cities, key=lambda x: self.cities[x].count, reverse=True)
        del cities_sorted[10:]
        self.print_for_cities("Доля вакансий по городам (в порядке убывания):", self.cities,
                              cities_sorted, "count")

    def print_for_cities(self, output: str, field: dict, names: list, value_name: str):
        """
        печать по городам
        :param output: str
            начало строки вывода
        :param field: dict
            Словарь, по которому нужно печатать статистику
        :param names:
            Названия городов
        :param value_name:
            название значения, которое нужно взять
        """
        flag = False
        print(output, end='')
        index = 0
        for name in names:
            if index == 0:
                print(' {', end='')
            printEnd = ', '
            if index == len(names) - 1:
                printEnd = ''
                flag = True
            print(f"'{name}': {getattr(field[name], value_name)}", end=printEnd)
            index += 1
        if flag:
            print('}')

    def print_one(self, output: str, field: dict, value_name: str):
        """
        Печатает одно свойство
        :param output: str
            начало строки вывода
        :param field: dict
            поле, которое нужно печатать
        :param value_name: str
            название значения, которое нужно взять
        """
        flag = False
        print(output, end='')
        index = 0
        for year in field.keys():
            if index == 0:
                print(' {', end='')
                index += 1
            printEnd = ', '
            if year == max(field.keys()):
                printEnd = ''
                flag = True
            print(f"{year}: {getattr(field[year], value_name)}", end=printEnd)
        if flag:
            print('}')


class Report:
    """
    Класс для генерации файлов статистики

    Attributes
    ----------
    years_salary: dict
        динамика зарплат вакансий по году
    years_count: dict
        динамика количества вакансий по году
    years_salary_vac: dict
        динамика зарплаты для вакансии по году
    years_count_vac: dict
        динамика количества для вакансии по году
    area_salary: dict
        уровень зарплат по городам
    area_count: dict
        доля вакансий по городам
    prof: dict
        название професии
    """
    years_salary = {}
    years_count = {}
    years_salary_vac = {}
    years_count_vac = {}
    area_salary = {}
    area_count = {}
    prof = ''

    def years_preparer(self, field: dict, value_name: str, dest: dict):
        """
        подготовка словаря лет
        """
        for year in field.keys():
            dest[year] = getattr(field[year], value_name)

    def citites_preparer(self, field: dict, names: list, value_name: str, dest: dict):
        """
        Подготовка словаря городов
        """
        for name in names:
            dest[name] = getattr(field[name], value_name)

    def prepare_data(self, years: dict, vacancies: dict, cities: dict, prof: str):
        """
        Подготовка словарей для статистики
        """
        self.years_preparer(years, "totalSalary", self.years_salary)
        self.years_preparer(years, "count", self.years_count)
        self.years_preparer(vacancies, "totalSalary", self.years_salary_vac)
        self.years_preparer(vacancies, "count", self.years_count_vac)

        cities_sorted = sorted(cities, key=lambda x: cities[x].totalSalary, reverse=True)
        del cities_sorted[10:]
        self.citites_preparer(cities, cities_sorted, "totalSalary", self.area_salary)

        cities_sorted = sorted(cities, key=lambda x: cities[x].count, reverse=True)
        del cities_sorted[10:]
        self.citites_preparer(cities, cities_sorted, "count", self.area_count)

    def as_text(self, value: str):
        if value is None:
            return ""
        return str(value)

    def generate_excel(self):
        """
        Генерация эксель таблицы
        """
        wb = Workbook()
        del wb['Sheet']
        sheet = wb.create_sheet('Статистика по годам')
        thin = Side(border_style="thin", color="000000")

        sheet["A1"] = "Год"
        sheet["A1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["A1"].font = Font(bold=True)

        sheet["B1"] = "Средняя зарплата"
        sheet["B1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["B1"].font = Font(bold=True)

        sheet["C1"] = "Средняя зарплата - Программист"
        sheet["C1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["C1"].font = Font(bold=True)

        sheet["D1"] = "Количество вакансий"
        sheet["D1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["D1"].font = Font(bold=True)

        sheet["E1"] = "Количество вакансий - Программист"
        sheet["E1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["E1"].font = Font(bold=True)

        for row, (year, value) in enumerate(self.years_salary.items(), start=2):
            sheet[f"A{row}"] = year
            sheet[f"A{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"B{row}"] = value
            sheet[f"B{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"C{row}"] = self.years_salary_vac[year]
            sheet[f"C{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"D{row}"] = self.years_count[year]
            sheet[f"D{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"E{row}"] = self.years_count_vac[year]
            sheet[f"E{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for column_cells in sheet.columns:
            length = max(len(self.as_text(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        sheet = wb.create_sheet('Статистика по городам')
        sheet["A1"] = "Город"
        sheet["A1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["A1"].font = Font(bold=True)

        sheet["B1"] = "Уровень зарплат"
        sheet["B1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["B1"].font = Font(bold=True)

        sheet["C1"].font = Font(bold=True)

        sheet["D1"] = "Город"
        sheet["D1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["D1"].font = Font(bold=True)

        sheet["E1"] = "Доля вакансий"
        sheet["E1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet["E1"].font = Font(bold=True)

        for row, (year, value) in enumerate(self.area_salary.items(), start=2):
            sheet[f"A{row}"] = year
            sheet[f"A{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"B{row}"] = value
            sheet[f"B{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for row, (year, value) in enumerate(self.area_count.items(), start=2):
            sheet[f"D{row}"] = year
            sheet[f"D{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[f"E{row}"] = value
            sheet[f"E{row}"].number_format = FORMAT_PERCENTAGE_00
            sheet[f"E{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for column_cells in sheet.columns:
            length = max(len(self.as_text(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        wb.save('report.xlsx')

    def generate_image(self):
        """
        Генерация картинки
        """
        figure, axes = pyplot.subplots(2, 2)
        w = 0.4
        X_axis = np.arange(len(self.years_salary.keys()))

        axes[0, 0].set_title('Уровень зарплат по годам')
        axes[0, 0].bar(X_axis - w / 2, self.years_salary.values(), width=w, label='средняя з/п')
        axes[0, 0].bar(X_axis + w / 2, self.years_salary_vac.values(), width=w, label='з/п программист')
        axes[0, 0].set_xticks(X_axis, self.years_salary.keys())
        axes[0, 0].set_xticklabels(self.years_salary.keys(), rotation='vertical', va='top', ha='center')
        axes[0, 0].grid(True, axis='y')
        axes[0, 0].tick_params(axis='both', labelsize=8)
        axes[0, 0].legend(fontsize=8)

        axes[0, 1].set_title('Количество вакансий по годам')
        axes[0, 1].bar(X_axis - w / 2, self.years_count.values(), width=w, label='Количество вакансий')
        axes[0, 1].bar(X_axis + w / 2, self.years_count_vac.values(), width=w, label='Количество вакансий\nпрограммист')
        axes[0, 1].set_xticks(X_axis, self.years_count.keys())
        axes[0, 1].set_xticklabels(self.years_count.keys(), rotation='vertical', va='top', ha='center')
        axes[0, 1].grid(True, axis='y')
        axes[0, 1].tick_params(axis='both', labelsize=8)
        axes[0, 1].legend(fontsize=8)

        axes[1, 0].set_title("Уровень зарплат по городам")
        axes[1, 0].invert_yaxis()
        salary = self.area_salary
        salaries = list(salary.keys())
        salaries = [label.replace(' ', '\n').replace('-', '-\n') for label in salaries]
        salary_values = list(salary.values())
        axes[1, 0].set_yticklabels(salaries, fontsize=6, va='center', ha='right')
        axes[1, 0].barh(salaries, salary_values)
        axes[1, 0].grid(True, axis='x')
        axes[1, 0].tick_params(axis='both', labelsize=8)

        axes[1, 1].set_title("Доля вакансий по городам")
        other = {'Другие': 1 - sum((list(self.area_count.values())))}
        other.update(self.area_count)
        self.area_count = other
        labels = list(self.area_count.keys())
        sizes = list(self.area_count.values())
        axes[1, 1].pie(sizes, labels=labels, textprops={'fontsize': 6})
        axes[1, 1].axis('scaled')

        pyplot.tight_layout()
        pyplot.savefig('graph.png', dpi=300)
        pyplot.show()

    def generate_pdf(self):
        """
        Генерация файла pdf
        """
        area_count_dic = self.area_count.items()
        area_count_dic = {x[0]: str(f'{x[1] * 100:,.2f}%').replace('.', ',') for x in area_count_dic}
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        header_year = ["Год", "Средняя зарплата", "Средняя зарплата - Программист", "Количество вакансий",
                       "Количество вакансий - Программист"]
        header_city = ["Город", "Уровень зарплат", "Город", "Доля вакансий"]
        pdf_template = template.render({'prof': self.prof,
                                        'years_salary_dic': self.years_salary,
                                        'years_count_dic': self.years_count,
                                        'years_salary_vac_dic': self.years_salary_vac,
                                        'years_count_vac_dic': self.years_count_vac,
                                        'area_salary_dic': self.area_salary,
                                        'area_count_dic': area_count_dic,
                                        'header_year': header_year,
                                        'header_city': header_city,
                                        'path': './graph.png'})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})


def get_statistic():
    """
    Получение статистики, генерация excel таблицы, картинки и pdf файла
    """
    inputer = InputConect()
    inputer.start_input()
    dataset = DataSet(inputer.file_name, list())
    dataset.fill_vacancies()
    inputer.count_vacancies(dataset.vacancies_objects)
    inputer.normalize_statistic()
    inputer.print_answer()
    reporter = Report()
    reporter.prepare_data(inputer.years, inputer.vacancies, inputer.cities, inputer.profession)
    reporter.generate_excel()
    reporter.generate_image()
    reporter.generate_pdf()
